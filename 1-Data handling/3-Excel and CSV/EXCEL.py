#https://openpyxl.readthedocs.io/en/stable/
from openpyxl import Workbook,load_workbook
from openpyxl.chart import BarChart,Reference

#to open excal
wb=load_workbook(filename='excl.xlsx')

ws=wb.active

for row in ws.iter_rows(min_row=1,min_col=1,max_row=4,max_col=5,values_only=True):
    print(row)

"""
 #create a sheet will be sheet number 0 by default and it is default name is Sheet


#to create more sheets

ws4=wb.create_sheet()#if we just make this the name will be Sheet1(as we make Sheet previus one) and the order of the sheet will be numbered Ascending

ws1=wb.create_sheet('MySheet1',0)

ws2=wb.create_sheet('MySheet2',1)#to make the sheet at beginning

ws3=wb.create_sheet('MySheet3',-1)#to make the sheet at end



#we can access the sheet by it is name

ws=wb['MySheet1']

ws['A4']=12 
#or
ws.cell(row=4,column=1,value=25)

wb.save('excl.xlsx')
"""

"""
data=[['ziad',11,22,33,44],
      ['ahmed',52,63,96,85],
      ['adel',45,12,36,52],
      ['ali',90,30,54,80]]


wb=Workbook()

ws=wb.active
for row in data:
    ws.append(row)

#make average function in rows
for i in range(1,len(data)+1):#1 to 4 range didn't use num 4
    ws.cell(row=i,column=6,value=f'=AVERAGE(B{i}:E{i})')
#make BarChart
values=Reference(ws,min_col=1,min_row=1,max_col=5,max_row=4)#the dimmention of the values
names=Reference(ws,min_col=1,min_row=1,max_col=1,max_row=4)#the dimention of names in first column

chart=BarChart()
chart.type='bar'
chart.title='Student degree'
chart.y_axis.title='Degree'
chart.x_axis.title='Student'
chart.width=20
chart.height=15
chart.legend=None
chart.add_data(values)
ws.add_chart(chart,'E15')#The chart is positioned on the worksheet at cell E15.


wb.save('excl.xlsx')

"""